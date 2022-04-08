from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
from buscarGrupo import buscarGrupo
from datetime import datetime
from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove
from salirGrupo import salirGrupo
import sesion as keepSession


#Se Carga el archivo Excel
excel = load_workbook('modulo2.xlsx')
#Selecciona la hoja dentro del excel
ws = excel['celular2']
#Selecciona el rango de la tabla
cell_range = ws['A2':'T500']

lst=[]

chrome = webdriver


# chrome = webdriver.Edge(executable_path='./msedgedriver')
# chrome.get('https://web.whatsapp.com/')

contador = -1

chrome = webdriver.Edge(executable_path='./msedgedriver')

wait = WebDriverWait(chrome,600)
chrome.implicitly_wait(20) # da una espera implícita de 20 segundos

chrome.get('https://web.whatsapp.com/')


for i in cell_range:
    
    time.sleep(2)
    contador += 1
    groupName = cell_range[contador][12].value
    condicion = cell_range[contador][17].value

    if(condicion != "RECICLAR" or condicion != "ELIMINADO"):
        
        if(groupName == "" or groupName == None):
                break

        if(cell_range[contador][17].value == "" or cell_range[contador][17].value == None):

            
                #Aqui empieza a Buscar el grupo
                x_nameFieldGroup = '//*[@id="side"]/div[1]/div/label/div/div[2]'
                nameFieldGroup = wait.until(ec.presence_of_element_located((By.XPATH,x_nameFieldGroup)))
                nameFieldGroup.clear()
                nameFieldGroup.send_keys(groupName)
                time.sleep(3)
                nameFieldGroup.send_keys(Keys.ENTER)
                try:
                    x_textoResultadoDeBusqueda = chrome.find_element_by_xpath('//*[@id="pane-side"]/div/div')
                    textoResultadoDeBusqueda = str(x_textoResultadoDeBusqueda.get_attribute("textContent"))

                    if(textoResultadoDeBusqueda == "No se encontró ningún chat, contacto ni mensaje"):
        
                        print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName)
                        cell_range[contador][17].value = "BUSCAR"
                        excel.save('modulo2.xlsx')


                    else:

                        #Click a Boton De 3 puntos
                        x_threePointsButton = '//*[@id="main"]/header/div[3]/div/div[2]/div/div'
                        buttonthreePoints = wait.until(ec.presence_of_element_located((By.XPATH,x_threePointsButton)))
                        buttonthreePoints.click()

                        #Click en Info Grupo
                        x_InfoGrupoButton = '//*[@id="app"]/div[1]/span[4]/div/ul/div/div/li[1]'
                        buttonInfoGrupo = wait.until(ec.presence_of_element_located((By.XPATH,x_InfoGrupoButton)))
                        buttonInfoGrupo.click()

                        time.sleep(1)
                        try:
                            extraeLaCantidadParticipantes = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[1]/div/div[3]/span/span/button')
                            extraerTexto = extraeLaCantidadParticipantes.get_attribute('textContent')
                            participantesExtraidos = int(extraerTexto[0] + extraerTexto[1])

                            if(participantesExtraidos > 5):

                                
                                try:

                                    #extrae el numero de participantes
                                    time.sleep(5)
                                    x_extraeLaTextoDeSalirDelGrupo = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[7]/div[1]/div[2]/div/span')
                                    extraeLaTextoDeSalirDelGrupo = x_extraeLaTextoDeSalirDelGrupo.get_attribute("textContent")
                                    extraeLaTextoDeSalirDelGrupo.strip()

                                except:

                                    #extrae el numero de participantes
                                    time.sleep(5)
                                    x_extraeLaTextoDeSalirDelGrupo = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[6]/div[1]/div[2]/div/span')
                                    extraeLaTextoDeSalirDelGrupo = x_extraeLaTextoDeSalirDelGrupo.get_attribute("textContent")
                                    extraeLaTextoDeSalirDelGrupo.strip()


                                salirGrupo(wait,groupName,chrome,cell_range, contador,excel,extraeLaTextoDeSalirDelGrupo)
                                print("Tenia ",participantesExtraidos," participantes")  

                            else:

                                print("El GRUPO: '", groupName,"' =>>> RECICLAR")
                                cell_range[contador][17].value = "RECICLAR"
                                excel.save('modulo2.xlsx')
                                print("Tiene: ",participantesExtraidos," participantes")  

                                    
                        except:
                            print("El GRUPO: '", groupName,"', ERROR")
                            cell_range[contador][17].value = "ERROR"
                            excel.save('modulo2.xlsx')
                        
                            
                except:
                    print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName)
                    cell_range[contador][17].value = "ELIMINADO"
                    excel.save('modulo2.xlsx')

