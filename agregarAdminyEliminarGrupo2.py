from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import time
from herramientas import agregarAdminyEliminarGrupo_tool
from herramientas import salirDelGrupo
from buscarGrupo import buscarGrupo
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from remove import remove
import sesion as keepSession


# Se Carga el archivo Excel
excel = load_workbook('modulo3_enlaces_actualizado.xlsx')
# Selecciona la hoja dentro del excel --3
ws = excel['celular2']
# Selecciona el rango de la tabla --3
cell_range = ws['A2':'AH300']

contador = -1


chrome = webdriver.Edge(executable_path='./msedgedriver')

driver = WebDriverWait(chrome,600)

chrome.get('https://web.whatsapp.com/')

action = ActionChains(chrome)


for i in cell_range:

    contador += 1

    if(cell_range[contador][19].value != None and cell_range[contador][27].value == None):

        if(i[0].value != None):

            groupName = cell_range[contador][19].value
            print("*******************************")
            x_nameFieldGroup = '//*[@id="side"]/div[1]/div/label/div/div[2]'
            nameFieldGroup = driver.until(
                ec.presence_of_element_located((By.XPATH, x_nameFieldGroup)))
            nameFieldGroup.clear()
            # da una espera implícita de 10 segundos
            chrome.implicitly_wait(10)
            nameFieldGroup.send_keys(groupName)
            nameFieldGroup.send_keys(Keys.ENTER)
            time.sleep(5)
            try:
                
                x_textoResultadoDeBusqueda = chrome.find_element(
                    By.XPATH, '//*[@id="pane-side"]/div/div')
                textoResultadoDeBusqueda = str(
                    x_textoResultadoDeBusqueda.get_attribute("textContent"))

                if(textoResultadoDeBusqueda == "No se encontró ningún chat, contacto ni mensaje"):

                    print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ", groupName)
                    cell_range[contador][27].value = "NO ENCONTRADO"
                    excel.save('modulo3_enlaces_actualizado.xlsx')
                else:
                    try:
                        x_extraeLaTextoDeSalirDelGrupo = driver.find_element(By.XPATH,'//*[@id="main"]/footer/div')
                        x_extraeLaTextoDeSalirDelGrupo = x_extraeLaTextoDeSalirDelGrupo.get_attribute("textContent")
                        extraeLaTextoDeSalirDelGrupo = x_extraeLaTextoDeSalirDelGrupo.strip()

                        if(extraeLaTextoDeSalirDelGrupo == "No puedes enviar mensajes porque ya no formas parte de este grupo."):
                            
                            #Click en Salir del grupo
                            x_botonEliminarGrupo = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[6]/div[1]/div[2]/div/span'
                            botonEliminarGrupo = driver.until(ec.presence_of_element_located((By.XPATH,x_botonEliminarGrupo)))
                            botonEliminarGrupo.click()

                            #Click  confirmar en Salir del grupo
                            x_botonConfirmarEliminacionDeGrupo = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div[2]/div[2]'
                            botonConfirmarEliminacionDeGrupo = driver.until(ec.presence_of_element_located((By.XPATH,x_botonConfirmarEliminacionDeGrupo)))
                            botonConfirmarEliminacionDeGrupo.click()

                            #Confirma el guardado de los datos en Excel
                            
                            print("Ya Contiene Administrador")
                            print("El Grupo: ", groupName)
                            cell_range[contador][27].value = "ELIMINADO"
                            excel.save('modulo3_enlaces_actualizado.xlsx')
                        else:
                            agregarAdminyEliminarGrupo_tool(driver,chrome,groupName,cell_range,contador,excel,action)

                    except:
                        agregarAdminyEliminarGrupo_tool(driver,chrome,groupName,cell_range,contador,excel,action)

# **********************************************************************MITAD DEL ALGORITMO******************************************************
            except:
                print("No Encontro el grupo")