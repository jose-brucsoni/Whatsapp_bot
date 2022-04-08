from tokenize import group
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
from buscarGrupo import buscarGrupo
from datetime import datetime
from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove
from salirGrupo import salirGrupo
import sesion as keepSession
from selenium.webdriver import ActionChains
from extraerDatos import extraerDatosDelExcel


#Se Carga el archivo Excel --2
excel = load_workbook('modulo2.xlsx')
#Selecciona la hoja dentro del excel --2
ws = excel['celular1']
#Selecciona el rango de la tabla --2
cell_range = ws['A2':'Z500']
#-----------------------------------------------------------
#Se Carga el archivo Excel --3
excel_3 = load_workbook('modulo3.xlsx')
#Selecciona la hoja dentro del excel --3
ws_3 = excel_3['modulo3 - Filtrado']
#Selecciona el rango de la tabla --3
cell_range_3 = ws_3['A2':'Z500']


#modificar en la funcion extraer datos, cambiarlo a excel modulo 3
(arregloNombres,arregloGrupo,arregloHorario,arregloSiglaMateria,arregloTurno,arregloModalidad,arregloMateriaSinAbreviar,arregloFilasDeExcel) = extraerDatosDelExcel()



chrome = webdriver


contador = -1
contador_2 = -1

#keepSession.newSession()
with open('./resource/session.log','r') as f:
    sesion=f.read().split(',')
    id=sesion[0]
    url=sesion[1]
    print(id+url)
chrome = keepSession.openSession(id,url)
wait = WebDriverWait(chrome,600)

action = ActionChains(chrome)

for i in cell_range:
    
        contador += 1
        if(cell_range[contador][12].value != None and cell_range[contador][18].value == None):
            contador_2 +=1
            groupName = cell_range[contador_2][12].value
            #obteniendo los datos
            nombreDelGrupo = arregloNombres[contador_2]
            grupo = arregloGrupo[contador_2]
            horario = arregloHorario[contador_2]
            siglaMateria = arregloSiglaMateria[contador]
            turno = arregloTurno[contador_2]
            modalidad = arregloModalidad[contador_2]
            materiaSinAbreviar = arregloMateriaSinAbreviar[contador_2]
            filaDeExcelAinsertar = arregloFilasDeExcel[contador_2]

            #Aqui empieza a Buscar el grupo 
            x_nameFieldGroup = '//*[@id="side"]/div[1]/div/label/div/div[2]'
            nameFieldGroup = wait.until(ec.presence_of_element_located((By.XPATH,x_nameFieldGroup)))
            nameFieldGroup.clear()
            nameFieldGroup.send_keys(groupName)
            time.sleep(3)
            nameFieldGroup.send_keys(Keys.ENTER)

            try:
                x_textoResultadoDeBusqueda = chrome.find_element_by_xpath('//*[@id="pane-side"]/div/div')
                textoResultadoDeBusqueda = str(x_textoResultadoDeBusqueda.get_attribute("textContent"))

                if(textoResultadoDeBusqueda == "No se encontró ningún chat, contacto ni mensaje"):

                    print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName)
                    cell_range[contador][17].value = "BUSCAR"
                    excel.save('modulo2.xlsx')


                else:
                    
                    #Click a Boton De 3 puntos
                    x_threePointsButton = '//*[@id="main"]/header/div[3]/div/div[2]/div/div'
                    buttonthreePoints = wait.until(ec.presence_of_element_located((By.XPATH,x_threePointsButton)))
                    buttonthreePoints.click()

                    #Click en Info Grupo
                    x_InfoGrupoButton = '//*[@id="app"]/div[1]/span[4]/div/ul/div/div/li[1]'
                    buttonInfoGrupo = wait.until(ec.presence_of_element_located((By.XPATH,x_InfoGrupoButton)))
                    buttonInfoGrupo.click()
                    time.sleep(2)

                    #extrae el numero de participantes
                    extraeLaCantidadParticipantes = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[1]/div/div[3]/span/span/button')
                    extraerTexto = extraeLaCantidadParticipantes.get_attribute('textContent')
                    participantesExtraidos = int(extraerTexto[0] + extraerTexto[1])

                    #extrae el admin
                    x_verificarAdmin = chrome.find_elements_by_xpath('//div[@data-testid="drawer-right"]//div[@class="_3uIPm WYyr1"]//div[@class="_3m_Xw"]')

                    for a in x_verificarAdmin:
                        usuarionombre= a.find_element_by_xpath('.//span[@class="ggj6brxn gfz4du6o r7fjleex g0rxnol2 lhj4utae le5p0ye3 l7jjieqr i0jNr"]').text

                        try:
                            admin = a.find_element_by_xpath('.//div[@class="_3bTNW"]').text
                        except:
                            admin="No es admin"

                        if usuarionombre=="Tú" and admin=="Admin. del grupo":
                            retorno = True
                            break
                        elif(usuarionombre =="Celular 2" and admin=="Admin. del grupo"):
                            cell_range[contador][21].value = "Celular 2 es Admin" #modificar de acuerdo al celular que se utilize
                            print("CELULAR 1 ES ADMIN")
                            excel.save('modulo2.xlsx')
                            excel_3.save('modulo3.xlsx')
                        
                        retorno = False
                        
                        
                        

                    if(participantesExtraidos <= 2 and retorno  == True):

                        #Click en el titulo del grupo para hacer aparecer el lapiz
                        x_clickEnLapizDeTitulo = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[1]/div/div[2]/div/div[1]/div/div/div[2]'
                        clickEnLapizDeTitulo = wait.until(ec.presence_of_element_located((By.XPATH,x_clickEnLapizDeTitulo)))
                        clickEnLapizDeTitulo.click()



                        
                        #Click en icono de lapiz de añadir titulo
                        x_botonLapizTitulo = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[1]/div/div[2]/div/div[1]/span[2]/button'
                        botonLapizTitulo = wait.until(ec.presence_of_element_located((By.XPATH,x_botonLapizTitulo)))
                        botonLapizTitulo.click()

                        
                        x_TextBoxTituloDeGrupo = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[1]/div/div[2]/div/div[1]/div/div/div[2]'
                        TextBoxTituloDeGrupo = wait.until(ec.presence_of_element_located((By.XPATH,x_TextBoxTituloDeGrupo)))
                        TextBoxTituloDeGrupo.clear()
                        TextBoxTituloDeGrupo.send_keys(nombreDelGrupo)
                        TextBoxTituloDeGrupo.send_keys(Keys.ENTER)

                        #Añadir Descripcion del Grupo
                        x_textoDeingresarDescripcion = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[2]/div[1]/div/div/div'
                        textoDeingresarDescripcion = wait.until(ec.presence_of_element_located((By.XPATH,x_textoDeingresarDescripcion)))
                        textoDeingresarDescripcion.click()

                        #Click en icono de lapiz de añadir descripcion
                        x_botonGuardarDescripcion = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[2]/div[1]/div/div/span[2]/button'
                        botonGuardarDescripcion = wait.until(ec.presence_of_element_located((By.XPATH,x_botonGuardarDescripcion)))
                        botonGuardarDescripcion.click()

                        
                        x_Bienvenida = "UTEPSA - Modulo(3)"
                        x_separador = "-----------------------"
                        x_materia = "Materia: " + materiaSinAbreviar
                        x_grupo = "Grupo: " + grupo
                        x_horario = "Horario: " + horario
                        x_turno = "Turno: " + turno
                        x_modalidad = "Modalidad: " + modalidad
                        x_siglaMateria = "Sigla de Materia: " + siglaMateria
                        x_docente = "Docente: "
                        x_Enlace = "Enlace: "

                        #ingresar descripcion y guardar
                        x_TextBoxIngresarDescripcion = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[2]/div[1]/div/div[1]/div/div/div[2]'
                        TextBoxIngresarDescripcion = wait.until(ec.presence_of_element_located((By.XPATH,x_TextBoxIngresarDescripcion)))
                        TextBoxIngresarDescripcion.clear()
                        TextBoxIngresarDescripcion.send_keys(x_Bienvenida+ (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT) + x_separador + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_materia + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_grupo + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_horario + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_turno + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_modalidad + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_siglaMateria + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_docente + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_Enlace)

                        TextBoxIngresarDescripcion.send_keys(Keys.ENTER)


                        #ingresar descripcion y guardar
                        x_textBoxDeChat = '//*[@id="main"]/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[2]'
                        textBoxDeChat = wait.until(ec.presence_of_element_located((By.XPATH,x_textBoxDeChat)))
                        textBoxDeChat.send_keys(x_Bienvenida+ (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT) + x_separador + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_materia + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_grupo + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_horario + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_turno + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_modalidad + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_siglaMateria + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_docente + (Keys.LEFT_SHIFT + Keys.ENTER) + (Keys.LEFT_SHIFT)+ x_Enlace)
                        textBoxDeChat.send_keys(Keys.ENTER)

                        
                        #Click Enlace de invitacion del grupo
                        x_EnlaceInvitacioButton = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[6]/div[2]/div[2]/div[2]'
                        buttonEnlaceInvitacio = wait.until(ec.presence_of_element_located((By.XPATH,x_EnlaceInvitacioButton)))
                        buttonEnlaceInvitacio.click()

                        
                        #Click en restablecer Enlace de invitacion del grupo
                        x_EnlaceRestablecerInvitacioButton = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/div[5]/div[2]/div/span'
                        buttonRestablecerEnlaceInvitacio = wait.until(ec.presence_of_element_located((By.XPATH,x_EnlaceRestablecerInvitacioButton)))
                        buttonRestablecerEnlaceInvitacio.click()

                        
                        #Click en Confirmar restablecer Enlace de invitacion del grupo
                        x_EnlaceConfirmarRestablecerInvitacioButton = '//*[@id="app"]/div[1]/span[2]/div[1]/span/div[1]/div/div/div/div/div[3]/div[2]/div/div'
                        buttonConfirmarRestablecerEnlaceInvitacio = wait.until(ec.presence_of_element_located((By.XPATH,x_EnlaceConfirmarRestablecerInvitacioButton)))
                        buttonConfirmarRestablecerEnlaceInvitacio.click()

                        time.sleep(3)

                        #Copia el Enlace de invitacion del grupo
                        link = chrome.find_element_by_xpath('//*[@id="group-invite-link-anchor"]')
                        retornarlink = link.get_attribute("href")

                        #Añadir codigo para insertar en el excel-------------
                        cell_range_3[contador_2][18].value = retornarlink
                        cell_range_3[contador_2][19].value = nombreDelGrupo
                        cell_range_3[contador_2][20].value = "Celular1" #modificar de acuerdo al celular que se utilize


                        #Añadir codigo para insertar en el excel-------------
                        cell_range[contador][17].value = "RECICLADO"
                        cell_range[contador][18].value = nombreDelGrupo
                        cell_range[contador][19].value = retornarlink
                        cell_range[contador][20].value = "Celular1"  #modificar de acuerdo al celular que se utilize

                        excel.save('modulo2.xlsx')
                        excel_3.save('modulo3.xlsx')  

                    

                    else:
                        print("******************---OBSERVACION---******************")
                        print("Grupo => ",groupName)
                        print("Contiene  => ",extraerTexto)
                        cell_range[contador_2][21].value = "CONTIENE MAS DE 2 INTEGRANTES"
                        excel.save('modulo2.xlsx')
                        excel_3.save('modulo3.xlsx')  
                        


                    
                

            except:
                print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName)
                cell_range[contador][17].value = "ERROR"
                excel.save('modulo2.xlsx')
                excel_3.save('modulo3.xlsx')


