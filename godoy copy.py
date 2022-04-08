from sys import breakpointhook
from traceback import print_tb
from openpyxl import load_workbook
import time

wb = load_workbook('modulo2.xlsx')
wb2 = load_workbook('modulo3_enlaces_actualizado.xlsx')

ws=wb['modulo2']
ws2=wb2['modulo3 - Filtrado']

cell = ws['A2':'AA500']
cell2 = ws2['A2':'AA500']

cont= -1
cont2= -1

for x in range(0,len(cell2)):

    cont2+=1



    for c in range (0,len(cell)):

        cont+=1
        

        #cell2[x][2].value==cell[c][1].value and cell2[x][3].value==cell[c][1].value and cell2[x][13].value==cell[c][5].value and cell2[x][14].value==cell[c][6].value
        if(cell[cont][1].value == cell2[cont2][2].value and cell[cont][3].value == cell2[cont2][3].value and cell[cont][4].value == cell2[cont2][7].value):

            
            cell[c][22].value = "Bimodular"
            cell2[x][23].value = "Bimodular"
            
            wb.save('modulo2.xlsx')
            wb2.save('modulo3_enlaces_actualizado.xlsx')
            break

        elif(cont == 498):
            
            cont = 0
            break
        elif(cont == 498 and cont2 == 431):

            exit()
