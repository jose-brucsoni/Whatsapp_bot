from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove

#Se Carga el archivo Excel --3
excel_3 = load_workbook('modulo3.xlsx')
#Selecciona la hoja dentro del excel --3
ws_3 = excel_3['modulo3 - Filtrado']
#Selecciona el rango de la tabla --3
cell_range = ws_3['A2':'Z500']


#Se Carga el archivo Excel --3
excel_plantilla = load_workbook('planilla_de_materias.xlsx')
#Selecciona la hoja dentro del excel --3
ws_plantilla = excel_plantilla['Hoja1']
#Selecciona el rango de la tabla --3
cell_range_de_la_plantilla = ws_3['A2':'Z500']


contador = -1
contador_de_planilla = -1
contador_buscador = -1



for i in cell_range_de_la_plantilla:

    contador_de_planilla +=1
    
    for i in cell_range:
            
            contador += 1

            if(cell_range[contador][3].value != None):

                if(i[0].value != None):

                    
                    modalidad = cell_range[contador_buscador][6].value 
                    turno = cell_range[contador_buscador][4].value 
                    horario = cell_range[contador_buscador][7].value 
                    grupo = cell_range[contador_buscador][2].value
                    materia = cell_range[contador][3].value
                    siglamateria = cell_range[contador][13].value
                    mat_codigo = cell_range[contador][2].value
                    tur_codigo = cell_range[contador][9].value
                    carrera = cell_range[contador][6].value


                    if(mat_codigo == cell_range_de_la_plantilla[contador_de_planilla][0].value and materia == cell_range_de_la_plantilla[contador_de_planilla][1].value and siglamateria == cell_range_de_la_plantilla[contador_de_planilla][2].value and tur_codigo == cell_range_de_la_plantilla[contador_de_planilla][3].value and carrera == cell_range_de_la_plantilla[contador_de_planilla][4].value ):
                        
                        for i in cell_range:
                            
                            contador_buscador +=1

                            materia_x = cell_range[contador_buscador][3].value
                            siglamateria_x = cell_range[contador_buscador][13].value
                            mat_codigo_x = cell_range[contador_buscador][2].value
                            tur_codigo_x = cell_range[contador_buscador][9].value
                            carrera_x = cell_range[contador_buscador][6].value
                            #------------------------------------------------
                            modalidad_x = cell_range[contador_buscador][6].value 
                            turno_x = cell_range[contador_buscador][4].value 
                            horario_x = cell_range[contador_buscador][7].value 
                            grupo_x = cell_range[contador_buscador][2].value
                            #---------------------------------------------------------
                            materia_Abuscar = cell_range_de_la_plantilla[contador_de_planilla][7].value
                            siglamateria_Abuscar = cell_range_de_la_plantilla[contador_de_planilla][8].value
                            mat_codigo_Abuscar = cell_range_de_la_plantilla[contador_de_planilla][6].value
                            tur_codigo_Abuscar = cell_range_de_la_plantilla[contador_de_planilla][9].value
                            carrera_Abuscar = cell_range_de_la_plantilla[contador_de_planilla][10].value

                            #añadir la condicion para buscar el nombre de la materia a buscar
                            if()
