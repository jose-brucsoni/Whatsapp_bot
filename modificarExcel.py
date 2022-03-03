from openpyxl  import Workbook
from openpyxl import load_workbook

#Se Carga el archivo Excel 1
excel = load_workbook('modulo2.xlsx')
#Selecciona la hoja dentro del excel
ws = excel['modulo2']
#Selecciona el rango de la tabla
cell_range = ws['A2':'N500']

#Se Carga el archivo Excel 2
excel2 = load_workbook('modulo2 v2.xlsx')
#Selecciona la hoja dentro del excel
ws2 = excel2['Hoja1']
#Selecciona el rango de la tabla
cell_range2 = ws2['A2':'O500']
filaPrueba = ws2['K2':'L500']

contadorExisten = 0
contadorNoExisten = 0
contadorDif = 0
contadorextra = 0


for x in range(499):

    nombreDeMateria = cell_range2[contadorDif][11].value
    grupoDeMateria = cell_range2[contadorDif][3].value
    #----------------------------
    nombreDeMateriaOriginal = cell_range[contadorextra][3].value
    grupoDeMateriaOriginal = cell_range[contadorextra][1].value

    if(nombreDeMateria == nombreDeMateriaOriginal and grupoDeMateria == grupoDeMateriaOriginal):
        
        contadorExisten +=1
        contadorDif +=1
        contadorextra = 0

    elif(x == 499):
        contadorNoExisten +=1

    elif(nombreDeMateria == None or nombreDeMateria == ""):
        break
    else:
        contadorextra +=1
    

print(contadorExisten)
print(contadorNoExisten)
print(x)