from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
from buscarGrupo import buscarGrupo
from datetime import datetime
from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove


chrome = webdriver.Chrome(executable_path='./chromedriver')

wait = WebDriverWait(chrome,600)
chrome.implicitly_wait(20) # da una espera implícita de 20 segundos



#Se Carga el archivo Excel
excel = load_workbook('modulo3_enlaces_actualizado.xlsx')
#Selecciona la hoja dentro del excel
ws = excel['modulo3 - Filtrado']
#Selecciona el rango de la tabla
cell_range = ws['A2':'Z500']


contador = -1

for i in cell_range:

        contador += 1

        if(cell_range[contador][18].value != None and cell_range[contador][19].value != None):

            enlace = str(cell_range[contador][18].value)
            nombreDeGrupo = str(cell_range[contador][19].value)
            chrome.get(enlace)
            
            
            nombreDegrupo = chrome.find_element_by_xpath('//*[@id="main_block"]/div[1]/h2')
            retornarnombreDegrupo = nombreDegrupo.get_attribute("textContent")


            if(retornarnombreDegrupo == nombreDeGrupo):

                cell_range[contador][21].value = "Verificado"
                print("Verificado----->",nombreDeGrupo)

            elif(retornarnombreDegrupo != nombreDeGrupo):
                
                cell_range[contador][21].value = "NO COINCIDE"
                cell_range[contador][22].value = retornarnombreDegrupo
                print("NO COINCIDE----->",retornarnombreDegrupo,"<-----No es igual a---> ",nombreDeGrupo)

            else:
                cell_range[contador][21].value = "ERROR"
                print("ERROR")


chrome.close()
excel.save('modulo3_enlaces_actualizado.xlsx')





