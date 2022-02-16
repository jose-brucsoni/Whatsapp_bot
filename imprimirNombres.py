from json import load
from typing import final
from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove

def extraerDatosDelExcel():
    
        
    print("Ingresa los siguientes Datos")
    nombreArchivoExtraido = input('Nombre del archivo Excel(obligatorio): ')
    hojaArchivo = input('Nombre de la hoja: ')
    nombreArchivoExcel = nombreArchivoExtraido + ".xlsx"

    #----------------------------------------------------------------------

    #Se Carga el archivo Excel
    excel = load_workbook(nombreArchivoExcel)
    #Selecciona la hoja dentro del excel
    ws = excel[hojaArchivo]
    #Selecciona el rango de la tabla
    cell_range = ws['A2':'K500']
    contador = -1
    arreglo = []




    for i in cell_range:
            contador += 1
            if(i[0].value != None):

                materia = cell_range[contador][3].value
                grupo = cell_range[contador][1].value
                nombreAbreviado = remove.abreviar(materia, grupo)
                print(nombreAbreviado)
                arreglo.append(nombreAbreviado)






                


    return arreglo