from traceback import print_tb
from openpyxl import load_workbook
import time

wb = load_workbook('mod3-actualizado.xlsx')
wb2 = load_workbook('modulo3_enlaces_actualizado.xlsx')
ws=wb['mod3']
ws2=wb2['modulo3 - Filtrado']
cell = ws['A2':'AA500']
cell2 = ws2['A2':'AA500']
cont=0
materiaFaltantes = 0

for x in range(0,len(cell2)):
    
    for c in range (0,len(cell)):
        #print(a)
        if(cell2[x][1].value==cell[c][1].value and cell2[x][2].value==cell[c][2].value and cell2[x][3].value==cell[c][3].value and cell2[x][6].value==cell[c][6].value and cell2[x][7].value==cell[c][7].value):
            cell[c][10].value = "Encontrado"
            cont+=1
        
            

wb.save('mod3-actualizado.xlsx')