from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
from buscarGrupo import buscarGrupo
from datetime import datetime
from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove

#----------------------------------------------------------------------
#Este proyecto convierte en administrador al contacto ya estando presente en el grupo, si ya es administrador el docente procedera a cerrarse y volver a comenzar el ciclo, tambien si hay 2 contactos registrado en el grupo el sistema retornara que "hay 2 docentes en el grupo"
#NOTA: Le falta que agregue los avisos en el excel


#Esto de aqui esta en espera de codigo para escribir lo que realize en el excel, por ejemplo: si ya tiene admins, si hay 2 docentes en el grupo o tambien sino tiene ningun docente registrado...
#Se Carga el archivo Excel
excel = load_workbook('numerosDocentes.xlsx')
#Selecciona la hoja dentro del excel
ws = excel['Hoja1']
#Selecciona el rango de la tabla
cell_range = ws['C4':'D200']


#Se Carga el archivo Excel 2
excel2 = load_workbook('modulo2.xlsx')
#Selecciona la hoja dentro del excel 2
ws2 = excel2['modulo2']
#Selecciona el rango de la tabla 2
cell_range2 = ws2['A2':'P500']

chrome = webdriver.Edge(executable_path='./msedgedriver')

wait = WebDriverWait(chrome,600)
chrome.implicitly_wait(20) # da una espera implícita de 20 segundos

chrome.get('https://web.whatsapp.com/')

groupName = "Grupo Para Pruebas"

time.sleep(2)


#Carga el nombre del grupo al textboxs
x_nameFieldGroup = '//*[@id="side"]/div[1]/div/label/div/div[2]'
nameFieldGroup = wait.until(ec.presence_of_element_located((By.XPATH,x_nameFieldGroup)))
nameFieldGroup.send_keys(groupName + Keys.ENTER)
chrome.implicitly_wait(20) # da una espera implícita de 20 segundos
x_encontrarElemento = '//*[@id="pane-side"]/div[1]/div/div/div[9]/div/div/div/div[2]/div[1]/div[1]/span'
x_encontrarElemento.__getattribute__("title")
time.sleep(2)

buscarGrupo.buscarGrupoEntextBox(x_encontrarElemento,groupName,nameFieldGroup,wait)
time.sleep(2)

#Click a Boton De 3 puntos
x_threePointsButton = '//*[@id="main"]/header/div[3]/div/div[2]/div/div'
buttonthreePoints = wait.until(ec.presence_of_element_located((By.XPATH,x_threePointsButton)))
buttonthreePoints.click()
time.sleep(2)

#Click en Info Grupo
x_InfoGrupoButton = '//*[@id="app"]/div[1]/span[4]/div/ul/div/div/li[1]'
buttonInfoGrupo = wait.until(ec.presence_of_element_located((By.XPATH,x_InfoGrupoButton)))
buttonInfoGrupo.click()

time.sleep(2)

#Click Lupa para buscar participante
x_lupaBuscarParticipante = '//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[6]/div[1]/div/div/div[2]/span'
lupaBuscarParticipante = wait.until(ec.presence_of_element_located((By.XPATH,x_lupaBuscarParticipante)))
lupaBuscarParticipante.click()

time.sleep(5)


try:

    try:
        x_VerificarSiEsAdmin = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div/div[2]/div[1]/div[2]')
        verificarSiEsAdmin = str(x_VerificarSiEsAdmin.get_attribute("textContent"))

        x_nombreDecontacto_1 = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div/div[2]/div[1]/div/span')
        nombreDecontacto_1 = str(x_nombreDecontacto_1.get_attribute("textContent"))

        x_nombreDecontacto_2 = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[3]/div/div/div[2]/div[1]/div/span')
        nombreDecontacto_2 = str(x_nombreDecontacto_2.get_attribute("textContent"))

        print(verificarSiEsAdmin,nombreDecontacto_1[0],nombreDecontacto_2[0],"-----------------------------------------")

        if(nombreDecontacto_1[0] != "+" and nombreDecontacto_2[0] == "+"):

            if(verificarSiEsAdmin != "Admin. del grupo" or verificarSiEsAdmin == None or verificarSiEsAdmin == ""):

                print("hay Conocidos---------------------------------")
                #click en el contacto, en el primero
                x_clickEnContacto = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div'
                clickEnContacto = wait.until(ec.presence_of_element_located((By.XPATH,x_clickEnContacto)))
                clickEnContacto.click()


                #Click en "Designar como admin del grupo"
                x_buttonDesignarAdministrador = '//*[@id="app"]/div[1]/span[4]/div/ul/div/li[1]/div[1]'
                buttonDesignarAdministrador = wait.until(ec.presence_of_element_located((By.XPATH,x_buttonDesignarAdministrador)))
                buttonDesignarAdministrador.click()


                #Salir de la ventana participantes
                x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
                salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
                salirDelaVentanaParticipantes.click()

            elif(verificarSiEsAdmin == "Admin. del grupo"):

                print("El Grupo: ", groupName, ", ya tiene administrador")
                #Salir de la ventana participantes
                x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
                salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
                salirDelaVentanaParticipantes.click()

        elif(nombreDecontacto_1[0] != "+" and nombreDecontacto_2[0] != "+"):

            print("----------------El Grupo: ", groupName, ", contiene 2 docentes-------------")
            #Salir de la ventana participantes
            x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
            salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
            salirDelaVentanaParticipantes.click()

        else:
            print("----------------El Grupo: ", groupName, ", NO TIENE NINGUN NUMERO REGISTRADO")
            #Salir de la ventana participantes
            x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
            salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
            salirDelaVentanaParticipantes.click()



    except:
        verificarSiEsAdmin = None
        x_nombreDecontacto_1 = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div/div[2]/div[1]/div/span')
        nombreDecontacto_1 = str(x_nombreDecontacto_1.get_attribute("textContent"))

        x_nombreDecontacto_2 = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[3]/div/div/div[2]/div[1]/div/span')
        nombreDecontacto_2 = str(x_nombreDecontacto_2.get_attribute("textContent"))

        print(verificarSiEsAdmin,nombreDecontacto_1[0],nombreDecontacto_2[0],"-----------------2------------------------")

        if(nombreDecontacto_1[0] != "+" and nombreDecontacto_2[0] == "+"):

            if(verificarSiEsAdmin != "Admin. del grupo" or verificarSiEsAdmin == None or verificarSiEsAdmin == ""):

                print("hay Conocidos---------------------------------")
                #click en el contacto, en el primero
                x_clickEnContacto = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/div[2]/div/div/div/div[4]/div/div'
                clickEnContacto = wait.until(ec.presence_of_element_located((By.XPATH,x_clickEnContacto)))
                clickEnContacto.click()


                #Click en "Designar como admin del grupo"
                x_buttonDesignarAdministrador = '//*[@id="app"]/div[1]/span[4]/div/ul/div/li[1]/div[1]'
                buttonDesignarAdministrador = wait.until(ec.presence_of_element_located((By.XPATH,x_buttonDesignarAdministrador)))
                buttonDesignarAdministrador.click()


                #Salir de la ventana participantes
                x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
                salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
                salirDelaVentanaParticipantes.click()

            elif(verificarSiEsAdmin == "Admin. del grupo"):

                print("El Grupo: ", groupName, ", ya tiene administrador")
                #Salir de la ventana participantes
                x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
                salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
                salirDelaVentanaParticipantes.click()

        elif(nombreDecontacto_1[0] != "+" and nombreDecontacto_2[0] != "+"):

            print("----------------El Grupo: ", groupName, ", contiene 2 docentes-------------")
            #Salir de la ventana participantes
            x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
            salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
            salirDelaVentanaParticipantes.click()

        else:
            print("----------------El Grupo: ", groupName, ", NO TIENE NINGUN NUMERO REGISTRADO")
            #Salir de la ventana participantes
            x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
            salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
            salirDelaVentanaParticipantes.click()


except:

    #Salir de la ventana participantes
    x_salirDelaVentanaParticipantes = '//*[@id="app"]/div[1]/span[2]/div[1]/div/div/div/div/div/div/header/div/div[1]/button/span'
    salirDelaVentanaParticipantes = wait.until(ec.presence_of_element_located((By.XPATH,x_salirDelaVentanaParticipantes)))
    salirDelaVentanaParticipantes.click()