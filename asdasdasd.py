from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
from buscarGrupo import buscarGrupo
from datetime import datetime
from openpyxl  import Workbook
from openpyxl import load_workbook
from remove import remove
from salirGrupo import salirGrupo
import sesion as keepSession


#Se Carga el archivo Excel
excel = load_workbook('modulo2.xlsx')
#Selecciona la hoja dentro del excel
ws = excel['modulo2']
#Selecciona el rango de la tabla
cell_range = ws['A2':'T500']

lst=[]

chrome = webdriver


# chrome = webdriver.Edge(executable_path='./msedgedriver')
# chrome.get('https://web.whatsapp.com/')

contador = -1

#keepSession.newSession()
with open('./resource/session.log','r') as f:
    sesion=f.read().split(',')
    id=sesion[0]
    url=sesion[1]
    print(id+url)
chrome = keepSession.openSession(id,url)
wait = WebDriverWait(chrome,600)

for i in cell_range:

    contador += 1
    groupName = cell_range[contador][12].value
    
    if(groupName == "" or groupName == None):
            break

    if(cell_range[contador][17].value == "" or cell_range[contador][17].value == None):

        
            #Aqui empieza a Buscar el grupo
            x_nameFieldGroup = '//*[@id="side"]/div[1]/div/label/div/div[2]'
            nameFieldGroup = wait.until(ec.presence_of_element_located((By.XPATH,x_nameFieldGroup)))
            nameFieldGroup.clear()
            nameFieldGroup.send_keys(groupName)
            time.sleep(5)
            nameFieldGroup.send_keys(Keys.ENTER)
            time.sleep(5)

            try:
                x_textoResultadoDeBusqueda = chrome.find_element_by_xpath('//*[@id="pane-side"]/div/div')
                textoResultadoDeBusqueda = str(x_textoResultadoDeBusqueda.get_attribute("textContent"))

                if(textoResultadoDeBusqueda == "No se encontró ningún chat, contacto ni mensaje"):
    
                    print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName)
                    cell_range[contador][17].value = "NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName
                    excel.save('modulo2.xlsx')
                
                print("NO SE ENCONTRO NINGUN GRUPO CON EL NOMBRE DE => ",groupName)
                cell_range[contador][17].value = "NO SE PUDO ENCONTRAR"
                excel.save('modulo2.xlsx')
                    
            except:
                
                time.sleep(1)

                #Click a Boton De 3 puntos
                x_threePointsButton = '//*[@id="main"]/header/div[3]/div/div[2]/div/div'
                buttonthreePoints = wait.until(ec.presence_of_element_located((By.XPATH,x_threePointsButton)))
                buttonthreePoints.click()
                time.sleep(1)

                #Click en Info Grupo
                x_InfoGrupoButton = '//*[@id="app"]/div[1]/span[4]/div/ul/div/div/li[1]'
                buttonInfoGrupo = wait.until(ec.presence_of_element_located((By.XPATH,x_InfoGrupoButton)))
                buttonInfoGrupo.click()

                time.sleep(1)
                try:

                    try:
                        #extrae el numero de participantes
                        x_extraeLaCantidadParticipantes = chrome.find_element_by_xpath('//*[@id="app"]/div[1]/div[1]/div[2]/div[3]/span/div[1]/span/div[1]/div/section/div[6]/div[1]/div/div/div[1]/span')
                        extraeLaCantidadParticipantes = str(x_extraeLaCantidadParticipantes.get_attribute("textContent"))

                        participantesExtraidos = extraeLaCantidadParticipantes[0] + extraeLaCantidadParticipantes[1]
                        print(participantesExtraidos)

                        if(participantesExtraidos > 5):
                            print("El GRUPO: '", groupName,"', SE PUDO BORRAR")
                            salirGrupo(wait,groupName,chrome,cell_range, contador,excel)

                   
                                
                    except:
                        print("El GRUPO: '", groupName,"', NO SE PUDO BORRAR_3")
                        cell_range[contador][17].value = "NO SE PUDO BORRAR_3"
                        excel.save('modulo2.xlsx')


                except:
                    
                    print("El GRUPO: '", groupName,"', NO SE PUDO BORRAR_4")
                    cell_range[contador][17].value = "NO SE PUDO BORRAR_4"
                    excel.save('modulo2.xlsx')
                    