from json import load
from typing import final
from attr import asdict
from openpyxl  import Workbook
from openpyxl import load_workbook
from herramientaParaEliminar import eliminarGrupoSeleccionado


    
#Se Carga el archivo Excel
excel = load_workbook('MATERIAS-1.xlsx')
#Selecciona la hoja dentro del excel
ws = excel['final']
#Selecciona el rango de la tabla
cell_range = ws['J2':'J39']

contador = -1
contador_para_foreach_2 = -1
arreglo = []


for i in cell_range:
        contador += 1
        if(i[0].value != None):
            #Se añade los datos al array devuelto
            arreglo.append(cell_range[contador][0].value)

for j in arreglo:

    contador_para_foreach_2 += 1
    nombre_del_arreglo = str(arreglo[contador_para_foreach_2])
    eliminarGrupoSeleccionado(nombre_del_arreglo)
    

    
    

