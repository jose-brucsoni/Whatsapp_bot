from openpyxl import load_workbook

wb = load_workbook('modulo2.xlsx')
wb2 = load_workbook('modulo3_enlaces_actualizado.xlsx')

ws=wb['modulo2']
ws2=wb2['modulo3 - Filtrado']

cell = ws['A2':'Z500']
cell2 = ws2['A2':'Z500']

cont= -1
cont2= -1
materiaFaltantes = 0
    
for i in cell:

    cont +=1
    print(cont,"0-1")
    print(cont2,"0-2")
    print("----------------------------")
    for j in cell2:

        cont2 +=1
        print(cont,"1-1")
        print(cont2,"1-2")
        # ----Grupo----------------------------------Nombre ---------------------------------------------- horario
        if(cell[cont][1].value == cell2[cont2][2].value and cell[cont][3].value == cell2[cont2][3].value and cell[cont][4].value == cell2[cont2][7].value ):
            
            cell[cont][22].value = "Bimodular"
            cell2[cont2][23].value = "Bimodular"
            print("Bimodular Encontrado")
            print("****************************************************")
            print(cont,"2-1")
            print(cont2,"2-2")
            break
        if(cont2 == 498):
            break
        
wb.save('modulo2.xlsx')
wb.save('modulo3_enlaces_actualizado.xlsx')