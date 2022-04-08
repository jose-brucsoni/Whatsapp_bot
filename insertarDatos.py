from openpyxl  import Workbook
from openpyxl import load_workbook


#Funcion para Insertar los datos en la hoja excel
def insertarDatosDelExcel(nombre,link,filadeExcelAinsertar):

    #Se Carga el archivo Excel --3
    wb = load_workbook('modulo3.xlsx')
    #Selecciona la hoja dentro del excel --3
    ws = wb['modulo3 - Filtrado']
    #Selecciona el rango de la tabla --3
    cell_range = ws['A2':'Z500']

    cell_range[filadeExcelAinsertar][18].value=link
    cell_range[filadeExcelAinsertar][19].value=nombre
    cell_range[filadeExcelAinsertar][20].value = "Celular1" #modificar de acuerdo al celular que se utilize

        
    #Confirma el guardado de los datos en Excel
    wb.save('modulo3.xlsx')
